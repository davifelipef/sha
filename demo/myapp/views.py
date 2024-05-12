import os
import io
import tempfile
from django.conf import settings
from django.db import transaction
from django.http import JsonResponse, HttpResponse, FileResponse
from django.shortcuts import render, HttpResponse, redirect
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.db.models import Q
from django.utils.cache import add_never_cache_headers
from .models import AtaResultados
from wsgiref.util import FileWrapper
from openpyxl import load_workbook
from xlsxwriter import Workbook

# Create your views here.

def home(request):
    return render(request, "index.html")

def validate_csv(uploaded_file):
    # Validar se o arquivo é um CSV
    if not uploaded_file.name.endswith('.csv'):
        raise ValidationError('Apenas arquivos CSV são permitidos.')

def importar_ata(request):
    if request.method == 'POST':
        # Obter o arquivo CSV enviado
        uploaded_file = request.FILES.get('filename')  

        # Validar o arquivo CSV (se enviado)
        if uploaded_file:
            try:
                validate_csv(uploaded_file)
                # Processar e salvar os dados no banco de dados
                processar_e_salvar_dados(uploaded_file) 
                # Exibir mensagem de confirmação
                messages.success(request, 'Ata importada com sucesso!')
                # Limpar o campo de envio de arquivo
                return redirect('importar_ata')
            except ValidationError as e:
                messages.error(request, e.message)
                return render(request, 'importar_ata.html')
    return render(request, 'importar_ata.html')

@transaction.atomic
def processar_e_salvar_dados(uploaded_file):
    # Dedodificação e leitura do arquivo enviado por upload
    decoded_file = uploaded_file.read().decode('utf-8').splitlines()

    # Declaração das variáveis ano, turma e série
    i_ano = ""
    i_turma = ""
    i_serie = ""

    for line_index, line in enumerate(decoded_file):  
        # Pula as linhas vazias
        if not line.strip():
            continue
        
        # Especifica que a separação dos dados é feita pelo ponto e vírgula
        row_data = line.split(';')

        # Check if the first field is empty, indicating an empty line
        if not row_data[0].strip():
            continue

        # Extração dos dados comuns a todos os alunos da ata
        # Extração do ano letivo
        if not i_ano:
            if 'Ano Letivo' in row_data:
                i_ano_index = row_data.index('Ano Letivo')
                i_ano = row_data[i_ano_index + 1].strip()
                print("Ano letivo extraído:", i_ano)
            else:
                i_ano = ""
        # Extração da turma
        if not i_turma:
            if 'Turma' in row_data:
                turma_index = row_data.index('Turma')
                i_turma = row_data[turma_index + 1].strip()
                print("Turma extraída:", i_turma)
                # Extração da série, para uso na etapa de confecção do histórico escolar
                i_serie = i_turma[:2].strip()
                print("Série extraída:", i_serie)
            else:
                i_turma = ""

        # Pula as linhas até a 12, onde estão os dados do estudante
        if line_index < 11:
            continue
        
        # Store student data only if it's not empty
        if row_data[2].strip():
            # Extrai o nome do estudante
            nome = row_data[2]
            # Extrai as notas do estudante
            notas = row_data[3:17]
            # Cria uma instância do banco de dados e atribui valores aos campos
            registro_estudante = AtaResultados(
                turma=i_turma,
                serie=i_serie,
                ano=i_ano,
                aluno=nome,
                nt_art=notas[0],
                nt_bio=notas[1],
                nt_edf=notas[2],
                nt_fil=notas[3],
                nt_fis=notas[4],
                nt_geo=notas[5],
                nt_his=notas[6],
                nt_lin=notas[7],
                nt_lpt=notas[8],
                nt_mat=notas[9],
                nt_pro=notas[10],
                nt_qui=notas[11],
                nt_soc=notas[12],
                nt_tec=notas[13]
            )
            # Salva os dados do estudante no banco de dados
            registro_estudante.save()

def emitir_historico(request):
    anos = AtaResultados.objects.values_list('ano', flat=True).distinct()
    turmas = AtaResultados.objects.values_list('turma', flat=True).distinct()

    if request.method == 'POST':
        print("Método POST acessado")
        selected_ano = request.POST.get('ano')

        if selected_ano:
            turmas = AtaResultados.objects.filter(ano=selected_ano).values('turma').distinct()
            print("Lista de turmas", turmas)
        else:
            turmas = "Selecionar Turma"  # Set turmas to empty list if no ano is selected
            alunos = "Selecionar Aluno"

        if request.POST.get('turma'):  # Check if turma is also selected in the form
            alunos = AtaResultados.objects.filter(turma=request.POST.get('turma')).values('aluno').distinct()
            return render(request, 'emitir_historico.html', {'anos': anos, 'selected_ano': selected_ano, 'turmas': turmas, 'alunos': alunos})
    elif request.method == 'GET':
        print("Método GET acessado")
        selected_ano = request.GET.get('ano')

        if selected_ano:
            print("Ano", selected_ano, "selecionado como filtro")
            turmas = AtaResultados.objects.filter(ano=selected_ano).values('turma').distinct()
            print("Turmas filtradas:", turmas)
            # Filter alunos based on turma (and potentially ano)
            alunos = AtaResultados.objects.filter(turma__in=turmas).values('aluno', 'nt_art', 'nt_bio', 'nt_edf', 'nt_fil', 'nt_fis', 'nt_geo', 'nt_his', 'nt_lin', 'nt_lpt', 'nt_mat', 'nt_pro', 'nt_qui', 'nt_soc', 'nt_tec')
            data = {'turmas': list(turmas), 'alunos': list(alunos)}
            #print("Dados recuperados:", data)
            return JsonResponse(data, safe=False)
        else:
            turmas = "Selecionar Turma"  # Set turmas to empty list if no ano is selected
            alunos = "Selecionar Aluno"

        return render(request, 'emitir_historico.html', {'anos': anos, 'selected_ano': selected_ano, 'turmas': turmas, 'alunos': alunos})

def upload_file(request):
  if request.method == 'POST':
    uploaded_file = request.FILES['file']
    # Validate file extension
    if not uploaded_file.name.endswith('.xlsx'):
      return render(request, 'upload.html', {'error': 'Formato de arquivo inválido. O formato aceito é .xlsx.'})
    # Save the file with a fixed name (overwrite if exists)
    filename = 'modelo_historico_em.xlsx'  # Replace with your desired name
    with open(os.path.join(settings.MEDIA_ROOT, filename), 'wb+') as destination:
      for chunk in uploaded_file.chunks():
        destination.write(chunk)
    return render(request, 'upload_file.html', {'success': 'Sucesso  no upload do arquivo!'})
  return render(request, 'upload_file.html')

# Definição do mapeamento notas/disciplinas
grade_fields = {  
    'grade_art': 'Arte',
    'grade_bio': 'Biologia',
    'grade_edf': 'Educação Física',
    'grade_fil': 'Filosofia',
    'grade_fis': 'Fisica',
    'grade_geo': 'Geografia',
    'grade_his': 'História',
    'grade_lin': 'Língua Estrangeira- Inglês',
    'grade_lpt': 'Língua Portuguesa e Literatura',
    'grade_mat': 'Matemática',
    'grade_pro': 'Projeto de Vida',
    'grade_qui': 'Quimica',
    'grade_soc': 'Sociologia',
    'grade_tec': 'Tecnologia e Inovação',
}

def get_subject_row(sheet):
  """
  This function iterates through rows in the sheet and returns the row number
  where the first cell contains a subject name from the `grade_fields` dictionary.
  """
  for row in sheet.iter_rows():
    if row[0].value and row[0].value in grade_fields.values():
      print("Notas atribuídas às disciplinas")
      return row[0].row
    else:
       print("Notas NÃO atribuídas às disciplinas")
  return None

def populate_and_download(request):
  if request.method == 'POST':
    print("Método de download acessado")
    selected_ano = request.POST.get('ano')
    print("Ano selecionado:", selected_ano)
    selected_turma = request.POST.get('turma')
    print("Turma selecionada:", selected_turma)
    selected_aluno = request.POST.get('aluno')
    print("Aluno selecionado:", selected_aluno)

    # Open uploaded Excel file (replace with actual filename and path)
    filename = 'modelo_historico_em.xlsx'  # Replace with your filename
    print("Modelo de histórico no excel aberto:", filename)
    wb = load_workbook(os.path.join(settings.MEDIA_ROOT, filename))
    sheet = wb.active  # Assuming you want to write to the active sheet

    # Place student information (access from request.POST if available)
    student_cell = sheet.cell(row=10, column=4)  # Replace with desired cell for student information
    student_cell.value = request.POST.get('aluno')  # Assuming student information is in a hidden field

    # Extract the first character (year indicator)
    year_indicator = int(selected_turma[:1])

    # Calculate target row and column for year data based on turma
    target_row = year_indicator + 13  # Adjust offset if needed (currently O15, Q15, or S15)
    print("Target row:", target_row)

    column_letters = ['O', 'Q', 'S']  # Include extra letters for flexibility

    # Adjust year indicator to match list index (considering offset)
    adjusted_index = year_indicator - 1

    # Check if adjusted index is within valid range
    if 0 <= adjusted_index < len(column_letters):
        target_col = column_letters[adjusted_index]
        print("Target col:", target_col)
    else:
        target_col = column_letters[adjusted_index]
        print("Target col missed:", target_col)
        print("Invalid year indicator in turma:", year_indicator)

    # Place grade data
    for html_element_id, grade_field in grade_fields.items():
      # Get grade value from the corresponding HTML element ID
      grade_value = request.POST.get(html_element_id)
      if grade_value:
        # Find subject row using the get_subject_row function
        subject_row = get_subject_row(sheet)

        if subject_row:
          subject_cell = sheet[f"{grade_field}{subject_row}"]
          # Check if subject cell exists and corresponding field name is available
          if subject_cell.value and subject_cell.value == grade_field:
            grade_cell = sheet.cell(row=subject_cell.row, column=ord(selected_turma[0]))  # Adjust cell based on turma
            grade_cell.value = grade_value

    # Save the modified Excel file
    modified_filename = 'modified_historico_em.xlsx'  # Replace with desired filename
    print("This is the modified filename:", modified_filename)
    modified_filepath = os.path.join(settings.MEDIA_ROOT, modified_filename)

    # Print the absolute path for verification
    print(f"Absolute path of saved file: {modified_filepath}")

    # Save directly to media root (remove temporary file creation)
    wb.save(modified_filepath)

    # Set up the response (assuming you still want to serve the file for download)
    #response = FileResponse(open(modified_filepath, 'rb'), as_attachment=True)

    response = HttpResponse(modified_filepath, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # Return the response
    print("Response:", response)
    return response

  # If the request method is not POST, return an empty HttpResponse
  return HttpResponse()